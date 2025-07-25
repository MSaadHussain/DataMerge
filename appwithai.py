import tkinter as tk
from tkinter import messagebox
import openpyxl
from rapidfuzz import process, fuzz
import json
import os
import re
import google.generativeai as genai
import time

# ========== CONFIG ==========
FILE1_PATH = "updated_benchmark.xlsx"
FILE2_PATH = "store_prices.xlsx"
PROGRESS_PATH = "progress.json"
GEMINI_API_KEY = "AIzaSyAxI9Qga0UjMVvV_b_AEwiW9uMzRBXVuxg"  # 🔑 Replace with your actual Gemini API key

# ========== SETUP GEMINI ==========
genai.configure(api_key=GEMINI_API_KEY)
gemini_model = genai.GenerativeModel("gemini-1.5-flash")

# ========== LOAD EXCEL ==========
file1_wb = openpyxl.load_workbook(FILE1_PATH)
file1_ws = file1_wb.active

file2_wb = openpyxl.load_workbook(FILE2_PATH)
file2_ws = file2_wb.active

file2_products = []
file2_data = {}

for row in file2_ws.iter_rows(min_row=2, values_only=True):
    name = row[0]
    prices = list(row[1:8])  # B to H (indices 0 to 6)
    if name:
        file2_products.append(name)
        file2_data[name] = prices

def load_progress():
    if os.path.exists(PROGRESS_PATH):
        with open(PROGRESS_PATH, "r") as f:
            return json.load(f).get("row", 2)
    return 2

def save_progress(row):
    with open(PROGRESS_PATH, "w") as f:
        json.dump({"row": row}, f)

class PriceMatcherApp:
    def __init__(self, master):
        self.master = master
        self.master.title("🧠 Product Matcher (Updated Columns)")
        self.master.geometry("1200x500")
        self.master.configure(bg="white")
        self.master.bind('<Return>', lambda e: self.update_and_next())

        self.current_row = load_progress()

        self.label = tk.Label(master, text="", font=("Arial", 16, "bold"), bg="white", anchor="w", justify="left")
        self.label.pack(fill="x", padx=20, pady=(20, 10))

        main_frame = tk.Frame(master, bg="white")
        main_frame.pack(pady=5, fill="both", expand=True)

        btn_frame = tk.Frame(main_frame, bg="white")
        btn_frame.pack(side="left", padx=20, fill="y")

        self.update_btn = tk.Button(btn_frame, text="✅ Update & Next (Enter)", command=self.update_and_next, width=25, font=("Arial", 13))
        self.update_btn.pack(pady=(0, 10))

        self.skip_btn = tk.Button(btn_frame, text="⏭️ Skip", command=self.next_product, width=25, font=("Arial", 13))
        self.skip_btn.pack()

        self.status = tk.Label(btn_frame, text="", font=("Arial", 11, "italic"), fg="gray", bg="white")
        self.status.pack(pady=10)

        self.text = tk.Text(main_frame, font=("Arial", 13), height=12, width=100, wrap="word", cursor="arrow")
        self.text.pack(side="right", fill="both", expand=True, padx=20)
        self.text.config(state=tk.DISABLED)
        self.text.bind("<Button-1>", self.on_click)

        self.best_matches = []
        self.load_product()

    def load_product(self):
        self.text.config(state=tk.NORMAL)
        self.text.delete(1.0, tk.END)
        self.best_matches = []

        item_cell = file1_ws[f"B{self.current_row}"]
        self.current_item_name = item_cell.value

        if not self.current_item_name:
            messagebox.showinfo("Done", f"Finished all rows up to {self.current_row}")
            return

        # 🛑 Skip if any price is already filled
        already_filled = any(file1_ws.cell(row=self.current_row, column=col).value not in [None, "", 0] for col in range(3, 9))
        if already_filled:
            self.current_row += 1
            self.load_product()
            return

        self.label.config(text=f"🔍 Searching: {self.current_item_name}")
        self.status.config(text=f"Row: {self.current_row}")

        words = re.findall(r'\b[a-zA-Z]+\b', self.current_item_name.lower())
        numbers = re.findall(r'\d+', self.current_item_name)

        filtered_candidates = []
        for name in file2_products:
            name_lc = name.lower()
            if all(num in name_lc for num in numbers) and any(word in name_lc for word in words):
                filtered_candidates.append(name)

        matches = process.extract(
            self.current_item_name,
            filtered_candidates,
            scorer=fuzz.partial_token_sort_ratio,
            limit=10,
            score_cutoff=0
        )

        for i, (match, score, _) in enumerate(matches):
            self.best_matches.append(match)
            words_in_line = match.split()
            for word in words_in_line:
                clean_word = re.sub(r'\W+', '', word).lower()
                tag = None
                if clean_word in words or clean_word in numbers:
                    tag = f"green_{i}"
                    self.text.insert(tk.END, word + " ", tag)
                    self.text.tag_config(tag, background="pale green")
                else:
                    self.text.insert(tk.END, word + " ")
            self.text.insert(tk.END, f"  ⟶ Score: {score}\n\n")

        if not self.best_matches:
            self.text.insert(tk.END, "❌ No good match found")
            time.sleep(0.5)  # Skip delay for bad match
            self.skip_btn.invoke()

        else:
            try:
                suggestion_prompt = (
                    f"You are helping to identify the best matching product from a list of store items "
                    f"based on a given benchmark product name.\n\n"
                    f"Benchmark Product Name:\n{self.current_item_name}\n\n"
                    f"Store Products:\n"
                    + "\n".join(f"{i+1}. {match}" for i, match in enumerate(self.best_matches))
                    + "\n\n"
                    f"❗Important Instructions:\n"
                    f"- Only one of these store products should match the benchmark name closely.\n"
                    f"- Packaging, usage, and purpose should be **identical or very close**.\n"
                    f"- Match should NOT just be based on matching words like 'LED' or 'white'.\n"
                    f"- Output ONLY the best match product name from the list — no extra text.if not match then none"
                )


                gemini_response = gemini_model.generate_content(suggestion_prompt)
                suggestion_text = gemini_response.text.strip()

                self.text.insert(tk.END, "\n🤖 Gemini Suggestion:\n", "gemini_title")
                self.text.insert(tk.END, suggestion_text + "\n", "gemini_text")
                self.text.tag_config("gemini_title", foreground="blue", font=("Arial", 13, "bold"))
                self.text.tag_config("gemini_text", foreground="darkblue", font=("Arial", 12, "italic"))

                # 🔵 Highlight Gemini's match in list
                for i, match in enumerate(self.best_matches):
                    if suggestion_text.lower().strip() == match.lower().strip():
                        line_start = f"{i * 3 + 1}.0"
                        line_end = f"{i * 3 + 3}.0"
                        self.text.tag_add("gemini_match", line_start, line_end)
                        self.text.tag_config("gemini_match", background="#DDEEFF", foreground="navy", font=("Arial", 12, "bold"))
                        break

            except Exception as e:
                self.text.insert(tk.END, f"\n⚠️ Gemini Error: {str(e)}", "error")
                self.text.tag_config("error", foreground="red")

        self.text.config(state=tk.DISABLED)

    def on_click(self, event):
        index = self.text.index(f"@{event.x},{event.y}")
        line = int(index.split('.')[0])
        match_idx = (line - 1) // 3
        if 0 <= match_idx < len(self.best_matches):
            self.selected_match_index = match_idx
            self.update_and_next()

    def update_and_next(self):
        match_idx = getattr(self, "selected_match_index", 0)
        if not self.best_matches:
            return

        selected = self.best_matches[match_idx]

        if selected in file2_data:
            prices = file2_data[selected]
            file1_ws.cell(row=self.current_row, column=3).value = prices[0]  # Carrefour
            merged_westzone = prices[4] if prices[4] else prices[6]
            file1_ws.cell(row=self.current_row, column=4).value = merged_westzone  # West Zone/Wmart
            file1_ws.cell(row=self.current_row, column=5).value = prices[3]  # Grandios
            file1_ws.cell(row=self.current_row, column=6).value = prices[1]  # Fresho
            file1_ws.cell(row=self.current_row, column=7).value = prices[2]  # Fresh N Rich
            file1_ws.cell(row=self.current_row, column=8).value = prices[5]  # Rayyan

            file1_wb.save(FILE1_PATH)
            save_progress(self.current_row + 1)

            self.current_row += 1
            self.selected_match_index = 0
            self.load_product()

    def next_product(self):
        save_progress(self.current_row + 1)
        self.current_row += 1
        self.selected_match_index = 0
        self.load_product()

root = tk.Tk()
app = PriceMatcherApp(root)
root.mainloop()
